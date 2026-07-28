#!/usr/bin/env python3
"""Import the five-sheet 三谋吕布 演武 workbook into database.json.

The command is a dry run unless ``--apply`` is supplied.  It parses every
source sheet, validates all catalog references before rendering, and replaces
the database atomically only when the rendered bytes actually changed.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import stat
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "三谋吕布-演武.xlsx"
DEFAULT_DATABASE = ROOT / "web/public/game-data/database.json"

EXPECTED_SHEETS = (
    "国家排行榜",
    "强队排行榜",
    "克制关系",
    "夺冠御三家",
    "阵容解析",
)
PROVIDER = "三谋吕布"
WORKBOOK_NAME = "三谋吕布-演武.xlsx"
UPDATED_AT = "2026-07-28"
ATTRIBUTION = "攻略数据由三谋吕布提供"

HERO_RANKINGS = ("S", "A", "B", "C", "D")
TEAM_RANKINGS = ("S", "A", "B")
RANKING_ORDER = {ranking: index for index, ranking in enumerate(HERO_RANKINGS)}
SECTION_TO_CAMP = {
    "魏国": "魏",
    "蜀国": "蜀",
    "吴国": "吴",
    "群雄": "群",
}
TEAM_SOURCES = ("strong", "championship")

HERO_ALIASES = {
    "sp诸葛亮": "诸葛亮2",
    "SP诸葛亮": "诸葛亮2",
    "sp周瑜": "周瑜2",
    "SP周瑜": "周瑜2",
    "皇甫嵩": "皇甫嵩2",
    "木鹿": "木鹿大王",
    "糜夫人": "麋夫人",
}

SKILL_ALIASES = {
    "万军": "万军辟易",
    "乘虚": "乘虚而入",
    "以静": "以静制动",
    "冲锐": "冲锐巧变",
    "势如": "势如破竹",
    "及锋": "及锋而试",
    "奇正": "奇正相生",
    "威名": "威名显赫",
    "御敌": "御敌临前",
    "忘私": "忘私相助",
    "惩前": "惩前毖后",
    "战八": "战八方",
    "折冲": "折冲御侮",
    "披坚": "披坚执锐",
    "拔刀": "拔刀相向",
    "拔刀相助": "拔刀相向",
    "指点": "指点乾坤",
    "挫锐": "挫锐折锋",
    "掠阵": "掠阵破军",
    "摧坚": "摧坚克难",
    "攻其": "攻其不备",
    "料事": "料事如神",
    "明起虚实": "明其虚实",
    "暗度阴平": "暗渡阴平",
    "横征": "横征暴敛",
    "横扫": "横扫千军",
    "百战": "百战不殆",
    "破军": "破军袭敌",
    "胜敌": "胜敌益强",
    "蓄势": "蓄势待发",
    "调和": "调和阴阳",
    "践墨": "践墨随敌",
    "运筹": "运筹帷幄",
    "金城": "金城汤池",
    "铁骑": "铁骑横冲",
    "铸甲": "铸甲销戈",
    "锐不": "锐不可当",
    "韬光": "韬光养晦",
}

FORMATION_ALIASES = {
    "方园阵": "方圆阵",
    "雁行阵": "雁形阵",
    "锥行阵": "锥形阵",
}
SQUARE_FORMATION_DESCRIPTION = "前后排协同布阵，兼顾攻防"

OUTCOME_BY_RGB = {
    "FF00B050": "largeAdvantage",
    "FFA9D18E": "smallAdvantage",
    "FF9DC3E6": "even",
    "FFFFD966": "smallDisadvantage",
    "FFFF5638": "largeDisadvantage",
    "FFA6A6A6": "self",
}
INVERSE_OUTCOME = {
    "largeAdvantage": "largeDisadvantage",
    "smallAdvantage": "smallDisadvantage",
    "even": "even",
    "smallDisadvantage": "smallAdvantage",
    "largeDisadvantage": "largeAdvantage",
    "self": "self",
}

# Matrix labels deliberately resolve to exact builds, not merely hero sets.
# The first two labels have the same heroes but different formations/skills.
MATCHUP_LABEL_ANCHORS = (
    ("司马懿+曹操+曹丕", "B8", ("司马懿", "曹操", "曹丕")),
    ("法刀司马+曹操+曹丕", "B12", ("司马懿", "曹操", "曹丕")),
    ("曹操+张春华+王异", "G16", ("曹操", "张春华", "王异")),
    ("姜维+sp诸葛亮+刘备", "B40", ("姜维", "诸葛亮2", "刘备")),
    ("sp周瑜+诸葛瑾+诸葛亮", "B65", ("周瑜2", "诸葛瑾", "诸葛亮")),
    ("祝融+孟获+sp诸葛亮", "G94", ("祝融", "孟获", "诸葛亮2")),
    ("袁术+皇甫嵩+孙坚", "B94", ("袁术", "皇甫嵩2", "孙坚2")),
    ("姜维+sp诸葛亮+黄月英", "G40", ("姜维", "诸葛亮2", "黄月英")),
    ("司马懿+郝昭+皇甫嵩", "G8", ("司马懿", "郝昭", "皇甫嵩2")),
    ("祝融+吴国太+貂蝉", "G98", ("祝融", "吴国太", "貂蝉")),
    ("孟获+祝融+木鹿", "B98", ("孟获", "祝融", "木鹿大王")),
    ("祝融+孟获+袁绍", "G102", ("祝融", "孟获", "袁绍")),
    ("袁术+皇甫嵩+朱儁", "L98", ("袁术", "皇甫嵩2", "朱儁")),
)


class ImportValidationError(ValueError):
    """Raised when workbook content cannot be imported exactly."""


@dataclass(frozen=True)
class ParsedBuild:
    ranking: str
    source: str
    section: str
    formation: str
    members: tuple[dict[str, Any], ...]
    origin: str


@dataclass(frozen=True)
class ImportStats:
    heroes: int
    skills: int
    strong_entries: int
    championship_entries: int
    teams: int
    cross_source_overlaps: int
    matchup_builds: int
    championship_groups: int
    analysis_sections: int
    analysis_points: int


EXPECTED_IMPORT_CARDINALITIES = {
    "strong_entries": 68,
    "championship_entries": 15,
    "teams": 77,
    "cross_source_overlaps": 3,
    "matchup_builds": 13,
    "championship_groups": 5,
    "analysis_sections": 2,
    "analysis_points": 6,
}


def _compact(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        return str(value).strip()
    return re.sub(r"\s+", "", value)


def _prose(value: Any) -> str:
    if not isinstance(value, str):
        raise ImportValidationError(f"expected text, got {value!r}")
    return re.sub(r"\s+", " ", value).strip()


def _require_mapping(value: Any, context: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ImportValidationError(f"{context} must be an object")
    return value


def normalize_hero(
    value: Any,
    hero_catalog: Mapping[str, Any],
    *,
    section: str | None = None,
    championship: bool = False,
    context: str = "hero",
) -> str:
    """Resolve one exact workbook hero token; never fuzzy-match."""

    raw = _compact(value)
    if not raw:
        raise ImportValidationError(f"{context}: empty hero")

    if raw == "孙坚" and (section == "群雄" or championship):
        hero = "孙坚2"
    else:
        hero = HERO_ALIASES.get(raw, raw)

    if hero not in hero_catalog:
        raise ImportValidationError(f"{context}: unknown hero {raw!r}")
    return hero


def normalize_skill(
    value: Any,
    skill_catalog: Mapping[str, Any],
    *,
    context: str = "skill",
) -> str:
    """Resolve one exact workbook skill token; never fuzzy-match."""

    raw = _compact(value)
    if not raw:
        raise ImportValidationError(f"{context}: empty skill")
    skill = SKILL_ALIASES.get(raw, raw)
    if skill not in skill_catalog:
        raise ImportValidationError(f"{context}: unknown skill {raw!r}")
    return skill


def parse_skill_slot(
    value: Any,
    skill_catalog: Mapping[str, Any],
    *,
    context: str,
) -> list[str]:
    if not isinstance(value, str) or not value.strip():
        raise ImportValidationError(f"{context}: skill slot must be non-empty text")
    alternatives: list[str] = []
    for token in re.split(r"\s*[/／]\s*", value.strip()):
        skill = normalize_skill(token, skill_catalog, context=context)
        if skill not in alternatives:
            alternatives.append(skill)
    if not alternatives:
        raise ImportValidationError(f"{context}: skill slot has no alternatives")
    return alternatives


def normalize_formation(
    value: Any,
    formations: Mapping[str, Any],
    *,
    context: str = "formation",
) -> str:
    raw = _compact(value)
    if not raw:
        raise ImportValidationError(f"{context}: empty formation")
    formation = FORMATION_ALIASES.get(raw, raw)
    if formation not in formations:
        raise ImportValidationError(f"{context}: unknown formation {raw!r}")
    return formation


def _validate_provider_and_date(workbook: Any) -> None:
    provider_cells = {
        "国家排行榜": "A2",
        "强队排行榜": "B2",
        "克制关系": "A1",
        "夺冠御三家": "A2",
        "阵容解析": "B2",
    }
    for sheet, coordinate in provider_cells.items():
        provider = _prose(workbook[sheet][coordinate].value)
        if provider != PROVIDER:
            raise ImportValidationError(
                f"{sheet}!{coordinate}: expected provider {PROVIDER!r}, got {provider!r}"
            )

    update_text = _compact(workbook["国家排行榜"]["G2"].value)
    match = re.fullmatch(r"更新[：:](\d{4})-(\d{1,2})-(\d{1,2})", update_text)
    if not match:
        raise ImportValidationError(
            "国家排行榜!G2: expected an 更新：YYYY-M-D source date"
        )
    normalized_date = f"{int(match[1]):04d}-{int(match[2]):02d}-{int(match[3]):02d}"
    if normalized_date != UPDATED_AT:
        raise ImportValidationError(
            f"国家排行榜!G2: expected source date {UPDATED_AT}, got {normalized_date}"
        )


def parse_hero_rankings(
    sheet: Worksheet,
    hero_catalog: Mapping[str, Any],
) -> dict[str, dict[str, str]]:
    current_section: str | None = None
    rankings: dict[str, dict[str, str]] = {}

    for row in range(1, sheet.max_row + 1):
        marker = _compact(sheet.cell(row, 1).value)
        if not marker:
            continue
        if marker in SECTION_TO_CAMP:
            current_section = marker
            continue
        if marker not in HERO_RANKINGS:
            if current_section is not None:
                raise ImportValidationError(
                    f"{sheet.title}!A{row}: unexpected ranking marker {marker!r}"
                )
            continue
        if current_section is None:
            raise ImportValidationError(
                f"{sheet.title}!A{row}: ranking appears before a country section"
            )

        heroes_in_row = 0
        for column in range(2, sheet.max_column + 1):
            raw = sheet.cell(row, column).value
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                continue
            coordinate = sheet.cell(row, column).coordinate
            hero = normalize_hero(
                raw,
                hero_catalog,
                section=current_section,
                context=f"{sheet.title}!{coordinate}",
            )
            if hero in rankings:
                raise ImportValidationError(
                    f"{sheet.title}!{coordinate}: duplicate ranking for {hero}"
                )
            rankings[hero] = {
                "ranking": marker,
                "camp": SECTION_TO_CAMP[current_section],
            }
            heroes_in_row += 1
        if heroes_in_row == 0:
            raise ImportValidationError(
                f"{sheet.title}!A{row}: ranking {marker} has no heroes"
            )

    missing = sorted(set(hero_catalog) - set(rankings))
    extra = sorted(set(rankings) - set(hero_catalog))
    if missing or extra:
        raise ImportValidationError(
            "国家排行榜 must rank every catalog hero exactly once; "
            f"missing={missing}, extra={extra}"
        )
    return rankings


def _parse_build(
    sheet: Worksheet,
    row: int,
    start_column: int,
    *,
    ranking: str,
    source: str,
    section: str,
    hero_catalog: Mapping[str, Any],
    skill_catalog: Mapping[str, Any],
    formations: Mapping[str, Any],
    championship: bool,
) -> ParsedBuild:
    origin_cell = sheet.cell(row, start_column)
    origin = f"{sheet.title}!{origin_cell.coordinate}"
    members: list[dict[str, Any]] = []
    for offset in range(3):
        hero_cell = sheet.cell(row, start_column + offset)
        hero = normalize_hero(
            hero_cell.value,
            hero_catalog,
            section=None if championship else section,
            championship=championship,
            context=f"{sheet.title}!{hero_cell.coordinate}",
        )
        slots = []
        for skill_row in (row + 1, row + 2):
            skill_cell = sheet.cell(skill_row, start_column + offset)
            slots.append(
                parse_skill_slot(
                    skill_cell.value,
                    skill_catalog,
                    context=f"{sheet.title}!{skill_cell.coordinate}",
                )
            )
        members.append({"hero": hero, "skillSlots": slots})

    formation_cell = sheet.cell(row, start_column + 3)
    formation = normalize_formation(
        formation_cell.value,
        formations,
        context=f"{sheet.title}!{formation_cell.coordinate}",
    )
    heroes = [member["hero"] for member in members]
    if len(set(heroes)) != 3:
        raise ImportValidationError(f"{origin}: a build must contain three unique heroes")

    return ParsedBuild(
        ranking=ranking,
        source=source,
        section=section,
        formation=formation,
        members=tuple(members),
        origin=origin,
    )


def parse_strong_builds(
    sheet: Worksheet,
    hero_catalog: Mapping[str, Any],
    skill_catalog: Mapping[str, Any],
    formations: Mapping[str, Any],
) -> list[ParsedBuild]:
    current_section: str | None = None
    current_ranking: str | None = None
    builds: list[ParsedBuild] = []

    for row in range(1, sheet.max_row + 1):
        marker = _compact(sheet.cell(row, 2).value)
        if marker in SECTION_TO_CAMP:
            current_section = marker
            current_ranking = None
        elif marker in TEAM_RANKINGS:
            if current_section is None:
                raise ImportValidationError(
                    f"{sheet.title}!B{row}: ranking appears before country section"
                )
            current_ranking = marker

        for start_column in (2, 7, 12):
            formation_value = sheet.cell(row, start_column + 3).value
            if formation_value is None or (
                isinstance(formation_value, str) and not formation_value.strip()
            ):
                continue
            if current_section is None or current_ranking is None:
                coordinate = sheet.cell(row, start_column + 3).coordinate
                raise ImportValidationError(
                    f"{sheet.title}!{coordinate}: build appears before section/ranking"
                )
            builds.append(
                _parse_build(
                    sheet,
                    row,
                    start_column,
                    ranking=current_ranking,
                    source="strong",
                    section=current_section,
                    hero_catalog=hero_catalog,
                    skill_catalog=skill_catalog,
                    formations=formations,
                    championship=False,
                )
            )

    if not builds:
        raise ImportValidationError("强队排行榜 contains no builds")
    return builds


def parse_championship_builds(
    sheet: Worksheet,
    hero_catalog: Mapping[str, Any],
    skill_catalog: Mapping[str, Any],
    formations: Mapping[str, Any],
) -> tuple[list[ParsedBuild], list[list[str]]]:
    current_ranking: str | None = None
    builds: list[ParsedBuild] = []
    group_origins: list[list[str]] = []

    for row in range(1, sheet.max_row + 1):
        marker = _compact(sheet.cell(row, 1).value)
        if marker in TEAM_RANKINGS:
            current_ranking = marker

        row_builds: list[ParsedBuild] = []
        for start_column in (1, 6, 11):
            formation_value = sheet.cell(row, start_column + 3).value
            if formation_value is None or (
                isinstance(formation_value, str) and not formation_value.strip()
            ):
                continue
            if current_ranking is None:
                coordinate = sheet.cell(row, start_column + 3).coordinate
                raise ImportValidationError(
                    f"{sheet.title}!{coordinate}: build appears before ranking"
                )
            row_builds.append(
                _parse_build(
                    sheet,
                    row,
                    start_column,
                    ranking=current_ranking,
                    source="championship",
                    section="夺冠御三家",
                    hero_catalog=hero_catalog,
                    skill_catalog=skill_catalog,
                    formations=formations,
                    championship=True,
                )
            )
        if row_builds:
            if len(row_builds) != 3:
                raise ImportValidationError(
                    f"{sheet.title}!{row}: championship group must contain three builds"
                )
            builds.extend(row_builds)
            group_origins.append([build.origin for build in row_builds])

    if not builds:
        raise ImportValidationError("夺冠御三家 contains no builds")
    return builds, group_origins


def _build_payload(build: ParsedBuild) -> dict[str, Any]:
    return {
        "formation": build.formation,
        "members": copy.deepcopy(list(build.members)),
    }


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def make_build_id(payload: Mapping[str, Any]) -> str:
    canonical = _canonical_json(payload)
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]
    members = payload.get("members")
    if not isinstance(members, Sequence):
        raise ImportValidationError("cannot create build id without members")
    hero_part = "-".join(str(member["hero"]) for member in members)
    return f"yanwu-{hero_part}-{digest}"


def merge_builds(
    strong_builds: Sequence[ParsedBuild],
    championship_builds: Sequence[ParsedBuild],
) -> tuple[list[dict[str, Any]], dict[str, str], int]:
    accumulators: dict[str, dict[str, Any]] = {}
    origins: dict[str, str] = {}
    ids_to_canonical: dict[str, str] = {}

    for build in (*strong_builds, *championship_builds):
        payload = _build_payload(build)
        canonical = _canonical_json(payload)
        build_id = make_build_id(payload)
        previous_canonical = ids_to_canonical.get(build_id)
        if previous_canonical is not None and previous_canonical != canonical:
            raise ImportValidationError(f"content-id collision for {build_id}")
        ids_to_canonical[build_id] = canonical

        accumulator = accumulators.get(canonical)
        if accumulator is None:
            accumulator = {
                "id": build_id,
                "ranking": build.ranking,
                "sources": [build.source],
                "section": build.section,
                "payload": payload,
            }
            accumulators[canonical] = accumulator
        else:
            if build.source not in accumulator["sources"]:
                accumulator["sources"].append(build.source)
            if RANKING_ORDER[build.ranking] < RANKING_ORDER[accumulator["ranking"]]:
                accumulator["ranking"] = build.ranking
            if build.source == "strong":
                existing_section = accumulator["section"]
                if existing_section not in ("夺冠御三家", build.section):
                    raise ImportValidationError(
                        f"{build.origin}: duplicate build has conflicting sections "
                        f"{existing_section!r} and {build.section!r}"
                    )
                accumulator["section"] = build.section

        if build.origin in origins and origins[build.origin] != build_id:
            raise ImportValidationError(f"duplicate source origin {build.origin}")
        origins[build.origin] = build_id

    teams: list[dict[str, Any]] = []
    for accumulator in accumulators.values():
        sources = sorted(
            accumulator["sources"],
            key=TEAM_SOURCES.index,
        )
        payload = accumulator["payload"]
        teams.append(
            {
                "id": accumulator["id"],
                "ranking": accumulator["ranking"],
                "sources": sources,
                "section": accumulator["section"],
                "formation": payload["formation"],
                "members": payload["members"],
            }
        )

    overlap_count = sum(
        1 for team in teams if team["sources"] == ["strong", "championship"]
    )
    return teams, origins, overlap_count


def _cell_outcome(cell: Cell) -> str:
    if cell.fill.fill_type != "solid" or cell.fill.fgColor.type != "rgb":
        raise ImportValidationError(
            f"{cell.parent.title}!{cell.coordinate}: matchup outcome needs a solid RGB fill"
        )
    rgb = cell.fill.fgColor.rgb
    outcome = OUTCOME_BY_RGB.get(rgb)
    if outcome is None:
        raise ImportValidationError(
            f"{cell.parent.title}!{cell.coordinate}: unknown outcome color {rgb!r}"
        )
    return outcome


def parse_matchups(
    sheet: Worksheet,
    origin_to_id: Mapping[str, str],
    teams: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    expected_labels = tuple(item[0] for item in MATCHUP_LABEL_ANCHORS)
    column_labels = tuple(_prose(sheet.cell(3, column).value) for column in range(2, 15))
    row_labels = tuple(_prose(sheet.cell(row, 1).value) for row in range(4, 17))
    if column_labels != expected_labels:
        raise ImportValidationError(
            f"克制关系 column labels changed: {column_labels!r}"
        )
    if row_labels != expected_labels:
        raise ImportValidationError(f"克制关系 row labels changed: {row_labels!r}")

    teams_by_id = {team["id"]: team for team in teams}
    build_ids: list[str] = []
    for label, anchor, expected_heroes in MATCHUP_LABEL_ANCHORS:
        origin = f"强队排行榜!{anchor}"
        build_id = origin_to_id.get(origin)
        if build_id is None:
            raise ImportValidationError(
                f"克制关系 label {label!r} requires missing build {origin}"
            )
        actual_heroes = tuple(
            member["hero"] for member in teams_by_id[build_id]["members"]
        )
        if sorted(actual_heroes) != sorted(expected_heroes):
            raise ImportValidationError(
                f"{origin}: matchup label {label!r} expected heroes "
                f"{expected_heroes!r}, got {actual_heroes!r}"
            )
        build_ids.append(build_id)

    outcomes = [
        [_cell_outcome(sheet.cell(row, column)) for column in range(2, 15)]
        for row in range(4, 17)
    ]
    for row in range(13):
        for column in range(13):
            outcome = outcomes[row][column]
            if (row == column) != (outcome == "self"):
                raise ImportValidationError(
                    f"克制关系!{sheet.cell(row + 4, column + 2).coordinate}: "
                    "self outcome must appear exactly on the diagonal"
                )
            inverse = outcomes[column][row]
            if INVERSE_OUTCOME[outcome] != inverse:
                raise ImportValidationError(
                    f"克制关系 matchup is not reciprocal at row {row + 1}, "
                    f"column {column + 1}: {outcome!r} vs {inverse!r}"
                )

    return {
        "orientation": "column-build-vs-row-build",
        "buildIds": build_ids,
        "outcomes": outcomes,
    }


def build_championship_groups(
    group_origins: Sequence[Sequence[str]],
    origin_to_id: Mapping[str, str],
) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    ids_to_canonical: dict[str, str] = {}
    for origins in group_origins:
        try:
            team_ids = [origin_to_id[origin] for origin in origins]
        except KeyError as error:
            raise ImportValidationError(
                f"championship group references missing origin {error.args[0]}"
            ) from error
        if len(team_ids) != 3 or len(set(team_ids)) != 3:
            raise ImportValidationError(
                "a championship group must reference three distinct build ids"
            )
        canonical = _canonical_json(team_ids)
        group_id = make_championship_group_id(team_ids)
        previous = ids_to_canonical.get(group_id)
        if previous is not None and previous != canonical:
            raise ImportValidationError(f"content-id collision for {group_id}")
        if previous == canonical:
            raise ImportValidationError("duplicate championship group")
        ids_to_canonical[group_id] = canonical
        groups.append({"id": group_id, "teamIds": team_ids})
    return groups


def make_championship_group_id(team_ids: Sequence[str]) -> str:
    canonical = _canonical_json(list(team_ids))
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]
    return f"championship-{digest}"


def parse_analysis_sections(
    sheet: Worksheet,
    hero_catalog: Mapping[str, Any],
) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    current_section: str | None = None
    current_analysis: dict[str, Any] | None = None

    for row in range(1, sheet.max_row + 1):
        raw_text = sheet.cell(row, 2).value
        if not isinstance(raw_text, str) or not raw_text.strip():
            continue
        text = _prose(raw_text)
        compact = _compact(raw_text)

        if compact in SECTION_TO_CAMP:
            current_section = compact
            current_analysis = None
            continue
        if current_section is None:
            # Provider/contact preface is intentionally not imported.
            continue

        if compact in hero_catalog or HERO_ALIASES.get(compact) in hero_catalog:
            subject = normalize_hero(
                compact,
                hero_catalog,
                section=current_section,
                context=f"{sheet.title}!B{row}",
            )
            current_analysis = {
                "section": current_section,
                "subject": subject,
                "points": [],
            }
            sections.append(current_analysis)
            continue

        if current_analysis is None:
            raise ImportValidationError(
                f"{sheet.title}!B{row}: analysis point appears before a subject"
            )

        number = sheet.cell(row, 1).value
        if isinstance(number, bool):
            number = None
        if isinstance(number, (int, float)):
            expected = len(current_analysis["points"]) + 1
            if number != expected:
                raise ImportValidationError(
                    f"{sheet.title}!A{row}: expected analysis point {expected}, got {number}"
                )
            current_analysis["points"].append(text)
        elif current_analysis["points"]:
            current_analysis["points"][-1] += text
        else:
            raise ImportValidationError(
                f"{sheet.title}!B{row}: continuation appears before the first point"
            )

    if not sections:
        raise ImportValidationError("阵容解析 contains no analysis sections")
    for section in sections:
        if not section["points"]:
            raise ImportValidationError(
                f"阵容解析 section {section['section']}/{section['subject']} has no points"
            )
    return sections


def validate_generated_database(database: Mapping[str, Any]) -> None:
    heroes = _require_mapping(database.get("heroes"), "heroes")
    skills = _require_mapping(database.get("skills"), "skills")
    formations = _require_mapping(database.get("formations"), "formations")
    teams = database.get("team")
    guide = _require_mapping(database.get("yanwuGuide"), "yanwuGuide")

    for hero_name, raw_hero in heroes.items():
        hero = _require_mapping(raw_hero, f"heroes[{hero_name!r}]")
        if "label" in hero or "rank" in hero:
            raise ImportValidationError(
                f"heroes[{hero_name!r}] still contains label/rank"
            )
        if hero.get("ranking") not in HERO_RANKINGS:
            raise ImportValidationError(
                f"heroes[{hero_name!r}].ranking must be S/A/B/C/D"
            )

    expected_camps = {
        "孙坚2": "群",
        "周仓": "蜀",
        "关平": "蜀",
        "徐盛": "吴",
    }
    for hero_name, expected_camp in expected_camps.items():
        if hero_name in heroes and heroes[hero_name].get("camp") != expected_camp:
            raise ImportValidationError(
                f"heroes[{hero_name!r}].camp must be {expected_camp}"
            )

    for skill_name, raw_skill in skills.items():
        skill = _require_mapping(raw_skill, f"skills[{skill_name!r}]")
        if "tier" in skill or "note" in skill:
            raise ImportValidationError(
                f"skills[{skill_name!r}] still contains tier/note"
            )

    if formations.get("方圆阵") != SQUARE_FORMATION_DESCRIPTION:
        raise ImportValidationError("formations must contain the canonical 方圆阵")
    if re.search(r"\d", formations["方圆阵"]):
        raise ImportValidationError("方圆阵 description must not contain numbers")

    if not isinstance(teams, list) or not teams:
        raise ImportValidationError("team must be a non-empty array")
    team_ids: set[str] = set()
    build_canonicals: set[str] = set()
    for index, raw_team in enumerate(teams):
        team = _require_mapping(raw_team, f"team[{index}]")
        expected_keys = {
            "id",
            "ranking",
            "sources",
            "section",
            "formation",
            "members",
        }
        if set(team) != expected_keys:
            raise ImportValidationError(
                f"team[{index}] has unexpected schema: {sorted(team)}"
            )
        team_id = team["id"]
        if not isinstance(team_id, str) or not team_id:
            raise ImportValidationError(f"team[{index}].id must be text")
        if team_id in team_ids:
            raise ImportValidationError(f"duplicate team id {team_id}")
        team_ids.add(team_id)
        if team["ranking"] not in TEAM_RANKINGS:
            raise ImportValidationError(f"team[{index}].ranking must be S/A/B")
        sources = team["sources"]
        if (
            not isinstance(sources, list)
            or not sources
            or any(source not in TEAM_SOURCES for source in sources)
            or sources != sorted(set(sources), key=TEAM_SOURCES.index)
        ):
            raise ImportValidationError(f"team[{index}].sources is invalid")
        section = team["section"]
        if not isinstance(section, str) or section not in (
            *SECTION_TO_CAMP,
            "夺冠御三家",
        ):
            raise ImportValidationError(f"team[{index}].section is invalid")
        if "strong" in sources and section == "夺冠御三家":
            raise ImportValidationError(
                f"team[{index}] has strong source without country section"
            )
        if "strong" not in sources and section != "夺冠御三家":
            raise ImportValidationError(
                f"team[{index}] championship-only section is invalid"
            )
        if team["formation"] not in formations:
            raise ImportValidationError(f"team[{index}] has unknown formation")
        members = team["members"]
        if not isinstance(members, list) or len(members) != 3:
            raise ImportValidationError(f"team[{index}] must have three members")
        member_heroes: list[str] = []
        for member_index, raw_member in enumerate(members):
            member = _require_mapping(
                raw_member, f"team[{index}].members[{member_index}]"
            )
            if set(member) != {"hero", "skillSlots"}:
                raise ImportValidationError(
                    f"team[{index}].members[{member_index}] has unexpected schema"
                )
            hero = member["hero"]
            if hero not in heroes:
                raise ImportValidationError(
                    f"team[{index}].members[{member_index}] has unknown hero {hero!r}"
                )
            member_heroes.append(hero)
            slots = member["skillSlots"]
            if not isinstance(slots, list) or len(slots) != 2:
                raise ImportValidationError(
                    f"team[{index}].members[{member_index}] needs two skill slots"
                )
            for slot_index, alternatives in enumerate(slots):
                if (
                    not isinstance(alternatives, list)
                    or not alternatives
                    or len(set(alternatives)) != len(alternatives)
                    or any(skill not in skills for skill in alternatives)
                ):
                    raise ImportValidationError(
                        f"team[{index}].members[{member_index}]."
                        f"skillSlots[{slot_index}] is invalid"
                    )
        if len(set(member_heroes)) != 3:
            raise ImportValidationError(f"team[{index}] repeats a hero")
        payload = {"formation": team["formation"], "members": members}
        canonical = _canonical_json(payload)
        if canonical in build_canonicals:
            raise ImportValidationError(f"team[{index}] duplicates another build")
        build_canonicals.add(canonical)
        if make_build_id(payload) != team_id:
            raise ImportValidationError(f"team[{index}] id is not content-derived")

    if guide.get("schemaVersion") != 1:
        raise ImportValidationError("yanwuGuide.schemaVersion must be 1")
    expected_source = {
        "provider": PROVIDER,
        "workbook": WORKBOOK_NAME,
        "updatedAt": UPDATED_AT,
        "attribution": ATTRIBUTION,
    }
    if guide.get("source") != expected_source:
        raise ImportValidationError("yanwuGuide.source metadata is invalid")

    matchups = _require_mapping(guide.get("matchups"), "yanwuGuide.matchups")
    if set(matchups) != {"orientation", "buildIds", "outcomes"}:
        raise ImportValidationError("yanwuGuide.matchups has unexpected schema")
    if matchups["orientation"] != "column-build-vs-row-build":
        raise ImportValidationError("yanwuGuide.matchups orientation is invalid")
    build_ids = matchups["buildIds"]
    outcomes = matchups["outcomes"]
    if (
        not isinstance(build_ids, list)
        or len(build_ids) != 13
        or len(set(build_ids)) != 13
        or any(build_id not in team_ids for build_id in build_ids)
    ):
        raise ImportValidationError("yanwuGuide.matchups.buildIds is invalid")
    if not isinstance(outcomes, list) or len(outcomes) != 13:
        raise ImportValidationError("yanwuGuide.matchups.outcomes must be 13x13")
    allowed_outcomes = set(OUTCOME_BY_RGB.values())
    for row, outcome_row in enumerate(outcomes):
        if not isinstance(outcome_row, list) or len(outcome_row) != 13:
            raise ImportValidationError("yanwuGuide.matchups.outcomes must be 13x13")
        for column, outcome in enumerate(outcome_row):
            if outcome not in allowed_outcomes:
                raise ImportValidationError("unknown matchup outcome")
            if (row == column) != (outcome == "self"):
                raise ImportValidationError("self outcomes must be diagonal")
            if INVERSE_OUTCOME[outcome] != outcomes[column][row]:
                raise ImportValidationError("matchup outcomes must be reciprocal")

    groups = guide.get("championshipGroups")
    if not isinstance(groups, list) or not groups:
        raise ImportValidationError("yanwuGuide.championshipGroups must be non-empty")
    group_ids: set[str] = set()
    for index, raw_group in enumerate(groups):
        group = _require_mapping(
            raw_group, f"yanwuGuide.championshipGroups[{index}]"
        )
        if set(group) != {"id", "teamIds"}:
            raise ImportValidationError(
                f"yanwuGuide.championshipGroups[{index}] has unexpected schema"
            )
        if (
            not isinstance(group["id"], str)
            or not group["id"].startswith("championship-")
            or group["id"] in group_ids
        ):
            raise ImportValidationError("championship group id is invalid")
        group_ids.add(group["id"])
        referenced_ids = group["teamIds"]
        if (
            not isinstance(referenced_ids, list)
            or len(referenced_ids) != 3
            or len(set(referenced_ids)) != 3
            or any(team_id not in team_ids for team_id in referenced_ids)
        ):
            raise ImportValidationError("championship group teamIds are invalid")
        if make_championship_group_id(referenced_ids) != group["id"]:
            raise ImportValidationError(
                "championship group id is not content-derived"
            )
        if any(
            "championship"
            not in next(team for team in teams if team["id"] == team_id)["sources"]
            for team_id in referenced_ids
        ):
            raise ImportValidationError(
                "championship group references a non-championship team"
            )

    analyses = guide.get("analysisSections")
    if not isinstance(analyses, list) or not analyses:
        raise ImportValidationError("yanwuGuide.analysisSections must be non-empty")
    for index, raw_analysis in enumerate(analyses):
        analysis = _require_mapping(
            raw_analysis, f"yanwuGuide.analysisSections[{index}]"
        )
        if set(analysis) != {"section", "subject", "points"}:
            raise ImportValidationError(
                f"yanwuGuide.analysisSections[{index}] has unexpected schema"
            )
        if analysis["section"] not in SECTION_TO_CAMP:
            raise ImportValidationError("analysis section is invalid")
        if analysis["subject"] not in heroes:
            raise ImportValidationError("analysis subject is unknown")
        if (
            not isinstance(analysis["points"], list)
            or not analysis["points"]
            or any(
                not isinstance(point, str) or not point.strip()
                for point in analysis["points"]
            )
        ):
            raise ImportValidationError("analysis points are invalid")

    serialized = _canonical_json(database)
    forbidden = ("VX", "850509047", "微信", "联系方式")
    leaked = [token for token in forbidden if token in serialized]
    if leaked:
        raise ImportValidationError(
            f"generated database contains forbidden contact data: {leaked}"
        )


def validate_import_cardinalities(stats: ImportStats) -> None:
    actual = {
        field: getattr(stats, field)
        for field in EXPECTED_IMPORT_CARDINALITIES
    }
    if actual != EXPECTED_IMPORT_CARDINALITIES:
        raise ImportValidationError(
            "workbook cardinalities changed; "
            f"expected={EXPECTED_IMPORT_CARDINALITIES}, actual={actual}"
        )


def build_database(
    original_database: Mapping[str, Any],
    workbook: Any,
) -> tuple[dict[str, Any], ImportStats]:
    if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
        raise ImportValidationError(
            f"workbook sheets must be exactly {EXPECTED_SHEETS!r}; "
            f"got {tuple(workbook.sheetnames)!r}"
        )
    _validate_provider_and_date(workbook)

    database = copy.deepcopy(dict(original_database))
    heroes = _require_mapping(database.get("heroes"), "heroes")
    skills = _require_mapping(database.get("skills"), "skills")
    formations = _require_mapping(database.get("formations"), "formations")
    if not heroes or not skills or not formations:
        raise ImportValidationError("heroes, skills, and formations must be non-empty")

    formations["方圆阵"] = SQUARE_FORMATION_DESCRIPTION
    hero_rankings = parse_hero_rankings(workbook["国家排行榜"], heroes)
    for hero_name, hero_data in heroes.items():
        hero = _require_mapping(hero_data, f"heroes[{hero_name!r}]")
        hero.pop("label", None)
        hero.pop("rank", None)
        hero["ranking"] = hero_rankings[hero_name]["ranking"]
        hero["camp"] = hero_rankings[hero_name]["camp"]

    for skill_name, skill_data in skills.items():
        skill = _require_mapping(skill_data, f"skills[{skill_name!r}]")
        skill.pop("tier", None)
        skill.pop("note", None)

    strong_builds = parse_strong_builds(
        workbook["强队排行榜"],
        heroes,
        skills,
        formations,
    )
    championship_builds, group_origins = parse_championship_builds(
        workbook["夺冠御三家"],
        heroes,
        skills,
        formations,
    )
    teams, origin_to_id, overlap_count = merge_builds(
        strong_builds,
        championship_builds,
    )
    matchups = parse_matchups(
        workbook["克制关系"],
        origin_to_id,
        teams,
    )
    championship_groups = build_championship_groups(
        group_origins,
        origin_to_id,
    )
    analysis_sections = parse_analysis_sections(
        workbook["阵容解析"],
        heroes,
    )

    database["team"] = teams
    database["yanwuGuide"] = {
        "schemaVersion": 1,
        "source": {
            "provider": PROVIDER,
            "workbook": WORKBOOK_NAME,
            "updatedAt": UPDATED_AT,
            "attribution": ATTRIBUTION,
        },
        "matchups": matchups,
        "championshipGroups": championship_groups,
        "analysisSections": analysis_sections,
    }
    validate_generated_database(database)

    stats = ImportStats(
        heroes=len(heroes),
        skills=len(skills),
        strong_entries=len(strong_builds),
        championship_entries=len(championship_builds),
        teams=len(teams),
        cross_source_overlaps=overlap_count,
        matchup_builds=len(matchups["buildIds"]),
        championship_groups=len(championship_groups),
        analysis_sections=len(analysis_sections),
        analysis_points=sum(
            len(section["points"]) for section in analysis_sections
        ),
    )
    validate_import_cardinalities(stats)
    return database, stats


def render_database(database: Mapping[str, Any]) -> bytes:
    return (
        json.dumps(database, ensure_ascii=False, indent=2) + "\n"
    ).encode("utf-8")


def write_database_if_changed(
    path: Path,
    rendered: bytes,
    *,
    apply: bool,
) -> bool:
    current = path.read_bytes()
    if current == rendered or not apply:
        return False

    mode = stat.S_IMODE(path.stat().st_mode)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(rendered)
            handle.flush()
            os.fsync(handle.fileno())
        os.chmod(temporary_path, mode)
        os.replace(temporary_path, path)
        try:
            directory_descriptor = os.open(path.parent, os.O_RDONLY)
        except OSError:
            directory_descriptor = None
        if directory_descriptor is not None:
            try:
                os.fsync(directory_descriptor)
            finally:
                os.close(directory_descriptor)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return True


def load_source_database(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise ImportValidationError(f"cannot read database {path}: {error}") from error
    return _require_mapping(value, "database")


def run_import(
    workbook_path: Path,
    database_path: Path,
    *,
    apply: bool,
) -> tuple[ImportStats, bool, bool]:
    if workbook_path.name != WORKBOOK_NAME:
        raise ImportValidationError(
            f"workbook filename must be {WORKBOOK_NAME!r}, got {workbook_path.name!r}"
        )
    if not workbook_path.is_file():
        raise ImportValidationError(f"workbook does not exist: {workbook_path}")
    if not database_path.is_file():
        raise ImportValidationError(f"database does not exist: {database_path}")

    original = load_source_database(database_path)
    try:
        workbook = load_workbook(
            workbook_path,
            data_only=False,
            read_only=False,
        )
    except Exception as error:  # openpyxl exposes several input-specific errors.
        raise ImportValidationError(
            f"cannot read workbook {workbook_path}: {error}"
        ) from error
    try:
        generated, stats = build_database(original, workbook)
    finally:
        workbook.close()

    rendered = render_database(generated)
    changed = database_path.read_bytes() != rendered
    written = write_database_if_changed(database_path, rendered, apply=apply)
    return stats, changed, written


def _print_summary(
    stats: ImportStats,
    *,
    apply: bool,
    changed: bool,
    written: bool,
) -> None:
    print(f"mode={'apply' if apply else 'dry-run'}")
    print(f"heroes={stats.heroes}")
    print(f"skills={stats.skills}")
    print(f"strong_entries={stats.strong_entries}")
    print(f"championship_entries={stats.championship_entries}")
    print(f"teams={stats.teams}")
    print(f"cross_source_overlaps={stats.cross_source_overlaps}")
    print(f"matchup_builds={stats.matchup_builds}")
    print(f"championship_groups={stats.championship_groups}")
    print(f"analysis_sections={stats.analysis_sections}")
    print(f"analysis_points={stats.analysis_points}")
    print(f"changes={'yes' if changed else 'no'}")
    print(f"written={'yes' if written else 'no'}")
    if not apply and changed:
        print("dry-run only; pass --apply to replace the database atomically")


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Validate and import 三谋吕布-演武.xlsx. "
            "The default is a no-write dry run."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"source workbook (default: {DEFAULT_WORKBOOK})",
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE,
        help=f"database JSON (default: {DEFAULT_DATABASE})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="atomically replace database.json after all validation passes",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    try:
        stats, changed, written = run_import(
            args.workbook.resolve(),
            args.database.resolve(),
            apply=args.apply,
        )
    except ImportValidationError as error:
        print(f"error: {error}", file=sys.stderr)
        return 1
    _print_summary(
        stats,
        apply=args.apply,
        changed=changed,
        written=written,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
