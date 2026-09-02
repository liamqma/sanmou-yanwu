from __future__ import annotations

import copy
import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from data.import_yanwu_workbook import (
    ATTRIBUTION,
    MATCHUP_LABEL_IDENTITIES,
    OUTCOME_BY_RGB,
    PROVIDER,
    SQUARE_FORMATION_DESCRIPTION,
    UPDATED_AT,
    WORKBOOK_NAME,
    EXPECTED_IMPORT_CARDINALITIES,
    EXPECTED_SHEETS,
    SOURCE_PROVIDER_CELLS,
    ImportStats,
    ImportValidationError,
    ParsedBuild,
    _cell_outcome,
    _validate_provider,
    build_championship_groups,
    make_build_id,
    merge_builds,
    normalize_formation,
    normalize_hero,
    normalize_skill,
    parse_analysis_sections,
    parse_championship_builds,
    parse_hero_rankings,
    parse_matchups,
    parse_skill_rankings,
    parse_skill_slot,
    parse_strong_builds,
    validate_generated_database,
    validate_import_cardinalities,
    write_database_if_changed,
)


def test_reviewed_author_markers_cover_the_renamed_workbook_layout() -> None:
    assert SOURCE_PROVIDER_CELLS == {
        ("武将Tier", "A2"): "但丁与你",
        ("战法Tier", "A2"): "飞将吕布",
        ("强队Tier", "B2"): "飞将吕布",
        ("克制关系", "A1"): "飞将吕布",
        ("夺冠御三家", "A2"): "飞将吕布",
        ("阵容解析", "B2"): "飞将吕布",
    }
    workbook = Workbook()
    workbook.active.title = EXPECTED_SHEETS[0]
    for sheet_name in EXPECTED_SHEETS[1:]:
        workbook.create_sheet(sheet_name)
    for (sheet_name, coordinate), marker in SOURCE_PROVIDER_CELLS.items():
        workbook[sheet_name][coordinate] = marker

    _validate_provider(workbook)

    workbook["强队Tier"]["B2"] = "相似但未审核的名称"
    with pytest.raises(ImportValidationError, match="强队Tier!B2"):
        _validate_provider(workbook)


def test_exact_aliases_normalize_without_fuzzy_matching() -> None:
    heroes = {
        "诸葛亮2": {},
        "周瑜2": {},
        "皇甫嵩2": {},
        "木鹿大王": {},
        "糜夫人": {},
        "孙坚": {},
        "孙坚2": {},
    }
    assert normalize_hero("sp诸葛亮", heroes) == "诸葛亮2"
    assert normalize_hero("sp周瑜", heroes) == "周瑜2"
    assert normalize_hero("皇甫嵩", heroes) == "皇甫嵩2"
    assert normalize_hero("木鹿", heroes) == "木鹿大王"
    assert normalize_hero("麋夫人", heroes) == "糜夫人"
    assert normalize_hero("孙坚", heroes, section="吴国") == "孙坚"
    assert normalize_hero("孙坚", heroes, section="群雄") == "孙坚2"
    assert normalize_hero("孙坚", heroes, championship=True) == "孙坚2"

    skills = {
        "拔刀相向": {},
        "锐不可当": {},
        "明其虚实": {},
        "暗渡阴平": {},
        "瞋目横矛": {},
        "谋而后动": {},
        "铸甲销戈": {},
    }
    assert normalize_skill("拔刀相助", skills) == "拔刀相向"
    assert normalize_skill("明起虚实", skills) == "明其虚实"
    assert normalize_skill("暗度阴平", skills) == "暗渡阴平"
    assert normalize_skill("暗度", skills) == "暗渡阴平"
    assert normalize_skill("瞋目", skills) == "瞋目横矛"
    assert normalize_skill("谋而", skills) == "谋而后动"
    assert normalize_skill("铸甲", skills) == "铸甲销戈"
    assert parse_skill_slot(
        "拔刀相助 / 锐不 / 拔刀",
        skills,
        context="test",
    ) == ["拔刀相向", "锐不可当"]

    formations = {
        "方圆阵": SQUARE_FORMATION_DESCRIPTION,
        "雁形阵": "x",
        "锥形阵": "x",
    }
    assert normalize_formation("方\n园阵", formations) == "方圆阵"
    assert normalize_formation("雁\n行阵", formations) == "雁形阵"
    assert normalize_formation("锥\n行阵", formations) == "锥形阵"

    with pytest.raises(ImportValidationError, match="unknown hero"):
        normalize_hero("诸葛", heroes)
    with pytest.raises(ImportValidationError, match="unknown skill"):
        normalize_skill("相助", skills)
    with pytest.raises(ImportValidationError, match="unknown formation"):
        normalize_formation("圆阵", formations)


def test_national_rankings_allow_only_the_reviewed_unranked_hero_and_disambiguate_sunjian() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "武将Tier"
    sheet["A4"] = "魏国"
    sheet["A5"] = "S"
    sheet["B5"] = "司马懿"
    sheet["A11"] = "蜀国"
    sheet["A12"] = "D"
    sheet["B12"] = "周仓"
    sheet["A18"] = "吴国"
    sheet["A19"] = "B"
    sheet["B19"] = "孙坚"
    sheet["C19"] = "徐盛"
    sheet["A25"] = "群雄"
    sheet["A26"] = "A"
    sheet["B26"] = "孙坚"

    result = parse_hero_rankings(
        sheet,
        {
            "司马懿": {},
            "周仓": {},
            "孙坚": {},
            "孙坚2": {},
            "徐盛": {},
            "小乔": {},
        },
    )

    assert result == {
        "司马懿": {"ranking": "S", "camp": "魏"},
        "周仓": {"ranking": "D", "camp": "蜀"},
        "孙坚": {"ranking": "B", "camp": "吴"},
        "徐盛": {"ranking": "B", "camp": "吴"},
        "孙坚2": {"ranking": "A", "camp": "群"},
    }
    assert "小乔" not in result

    with pytest.raises(ImportValidationError, match="exact reviewed hero subset"):
        parse_hero_rankings(
            sheet,
            {
                "司马懿": {},
                "周仓": {},
                "孙坚": {},
                "孙坚2": {},
                "徐盛": {},
                "小乔": {},
                "大乔": {},
            },
        )


def test_skill_rankings_import_exact_categories_and_allow_unranked_catalog() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "战法Tier"
    catalog = {f"{category}战法": {} for category in ("兵刃", "谋略", "治疗", "防御", "辅助", "文武")}
    catalog["未排名战法"] = {}

    row = 4
    for ranking, category in zip("SABCDS", ("兵刃", "谋略", "治疗", "防御", "辅助", "文武")):
        sheet.cell(row, 1).value = category
        sheet.cell(row + 1, 1).value = ranking
        sheet.cell(row + 1, 2).value = f"{category}战法"
        row += 3

    result = parse_skill_rankings(sheet, catalog)

    assert result == {
        f"{category}战法": {"ranking": ranking, "category": category}
        for ranking, category in zip("SABCDS", ("兵刃", "谋略", "治疗", "防御", "辅助", "文武"))
    }
    assert "未排名战法" not in result


def test_championship_builds_are_s_rank_and_exact_duplicates_share_ids() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "夺冠御三家"
    heroes = {name: {} for name in ("甲", "乙", "丙", "丁", "戊", "己", "庚", "辛", "壬")}
    skills = {f"{hero}{slot}": {} for hero in heroes for slot in (1, 2)}
    skills["甲2-备选"] = {}
    for row in (5, 9):
        for start, names in zip((1, 6, 11), (("甲", "乙", "丙"), ("丁", "戊", "己"), ("庚", "辛", "壬"))):
            for offset, hero in enumerate(names):
                sheet.cell(row, start + offset).value = hero
                sheet.cell(row + 1, start + offset).value = f"{hero}1"
                sheet.cell(row + 2, start + offset).value = f"{hero}2"
            sheet.cell(row, start + 3).value = "方圆阵"
    sheet["A11"] = "甲2-备选"

    builds, group_origins = parse_championship_builds(
        sheet,
        heroes,
        skills,
        {"方圆阵": SQUARE_FORMATION_DESCRIPTION},
    )
    teams, origin_to_id, _ = merge_builds([], builds)
    groups = build_championship_groups(group_origins, origin_to_id)

    assert len(builds) == 6
    assert all(build.ranking == "S" for build in builds)
    assert len(teams) == 4
    assert len(groups) == 2
    assert groups[0]["teamIds"][0] != groups[1]["teamIds"][0]
    assert groups[0]["teamIds"][1:] == groups[1]["teamIds"][1:]


def test_strong_build_parser_normalizes_slots_and_formation() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "强队Tier"
    sheet["B5"] = "魏国"
    sheet["B7"] = "S"
    sheet["B8"] = "sp诸葛亮"
    sheet["C8"] = "皇甫嵩"
    sheet["D8"] = "木鹿"
    sheet["E8"] = "方\n园阵"
    sheet["B9"] = "拔刀相助 / 锐不"
    sheet["C9"] = "明起虚实"
    sheet["D9"] = "暗度阴平"
    sheet["B10"] = "铸甲"
    sheet["C10"] = "战八"
    sheet["D10"] = "金城"

    builds = parse_strong_builds(
        sheet,
        {"诸葛亮2": {}, "皇甫嵩2": {}, "木鹿大王": {}},
        {
            "拔刀相向": {},
            "锐不可当": {},
            "明其虚实": {},
            "暗渡阴平": {},
            "铸甲销戈": {},
            "战八方": {},
            "金城汤池": {},
        },
        {"方圆阵": SQUARE_FORMATION_DESCRIPTION},
    )

    assert len(builds) == 1
    assert builds[0].origin == "强队Tier!B8"
    assert builds[0].formation == "方圆阵"
    assert builds[0].members == (
        {
            "hero": "诸葛亮2",
            "skillSlots": [
                ["拔刀相向", "锐不可当"],
                ["铸甲销戈"],
            ],
        },
        {
            "hero": "皇甫嵩2",
            "skillSlots": [["明其虚实"], ["战八方"]],
        },
        {
            "hero": "木鹿大王",
            "skillSlots": [["暗渡阴平"], ["金城汤池"]],
        },
    )


def _parsed_build(
    *,
    origin: str,
    source: str,
    section: str,
    second_skill: str = "乙技二",
) -> ParsedBuild:
    return ParsedBuild(
        ranking="S",
        source=source,
        section=section,
        formation="方圆阵",
        members=(
            {"hero": "甲", "skillSlots": [["甲技一"], ["甲技二"]]},
            {"hero": "乙", "skillSlots": [["乙技一"], [second_skill]]},
            {"hero": "丙", "skillSlots": [["丙技一"], ["丙技二"]]},
        ),
        origin=origin,
    )


def test_exact_cross_source_builds_reuse_content_id_and_country_section() -> None:
    strong = _parsed_build(
        origin="强队排行榜!B8",
        source="strong",
        section="魏国",
    )
    duplicate_champion = _parsed_build(
        origin="夺冠御三家!A5",
        source="championship",
        section="夺冠御三家",
    )
    champion_only = _parsed_build(
        origin="夺冠御三家!F5",
        source="championship",
        section="夺冠御三家",
        second_skill="乙技三",
    )

    teams, origins, overlap_count = merge_builds(
        [strong],
        [duplicate_champion, champion_only],
    )

    assert len(teams) == 2
    assert overlap_count == 1
    overlap = teams[0]
    assert overlap["sources"] == ["strong", "championship"]
    assert overlap["section"] == "魏国"
    assert origins["强队排行榜!B8"] == origins["夺冠御三家!A5"]
    assert teams[1]["sources"] == ["championship"]
    assert teams[1]["section"] == "夺冠御三家"
    assert overlap["id"] == make_build_id(
        {"formation": overlap["formation"], "members": overlap["members"]}
    )


@pytest.mark.parametrize(
    ("rgb", "expected"),
    list(OUTCOME_BY_RGB.items()),
)
def test_matchup_fill_colors_parse_to_exact_outcome_enum(
    rgb: str,
    expected: str,
) -> None:
    workbook = Workbook()
    cell = workbook.active["A1"]
    cell.fill = PatternFill(fill_type="solid", fgColor=rgb)
    assert _cell_outcome(cell) == expected


def test_matchups_resolve_stable_build_identity_without_row_anchors() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "克制关系"
    teams = []
    expected_ids = []
    for index, (label, heroes, variant) in enumerate(MATCHUP_LABEL_IDENTITIES):
        sheet.cell(3, index + 2).value = label
        sheet.cell(index + 4, 1).value = label
        members = []
        for hero in heroes:
            if hero == "司马懿" and variant == "fadao":
                slots = [["运智铺谋"], ["谋而后动"]]
            elif hero == "司马懿":
                slots = [["未雨绸缪"], ["潜龙在渊"]]
            else:
                slots = [[f"{hero}战法一"], [f"{hero}战法二-{index}"]]
            members.append({"hero": hero, "skillSlots": slots})
        payload = {"formation": "方圆阵", "members": members}
        team_id = make_build_id(payload)
        expected_ids.append(team_id)
        teams.append(
            {
                "id": team_id,
                "ranking": "S",
                "sources": ["strong"],
                "section": "魏国",
                **payload,
            }
        )

    even_rgb = next(rgb for rgb, outcome in OUTCOME_BY_RGB.items() if outcome == "even")
    self_rgb = next(rgb for rgb, outcome in OUTCOME_BY_RGB.items() if outcome == "self")
    for row in range(13):
        for column in range(13):
            sheet.cell(row + 4, column + 2).fill = PatternFill(
                fill_type="solid",
                fgColor=self_rgb if row == column else even_rgb,
            )

    result = parse_matchups(sheet, teams)

    assert result["buildIds"] == expected_ids


def test_analysis_parser_joins_continuations_and_omits_contact_preface() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "阵容解析"
    sheet["B2"] = PROVIDER
    sheet["B3"] = "演武群，VX：850509047"
    sheet["B5"] = "魏国"
    sheet["B7"] = "司马懿"
    sheet["A8"] = 1
    sheet["B8"] = "第一段；"
    sheet["B9"] = "续段"
    sheet["A10"] = 2
    sheet["B10"] = "第二点"

    result = parse_analysis_sections(sheet, {"司马懿": {}})

    assert result == [
        {
            "section": "魏国",
            "subject": "司马懿",
            "points": ["第一段；续段", "第二点"],
        }
    ]
    assert "VX" not in json.dumps(result, ensure_ascii=False)
    assert "850509047" not in json.dumps(result, ensure_ascii=False)


def _valid_database() -> dict[str, object]:
    heroes = {
        hero: {"camp": "魏", "ranking": "S"}
        for hero in ("甲", "乙", "丙")
    }
    skills: dict[str, dict[str, object]] = {}
    teams = []
    for index in range(13):
        skill_names = [
            f"甲技{index}一",
            f"甲技{index}二",
            "乙技一",
            "乙技二",
            "丙技一",
            "丙技二",
        ]
        skills.update({name: {} for name in skill_names})
        members = [
            {
                "hero": "甲",
                "skillSlots": [[skill_names[0]], [skill_names[1]]],
            },
            {"hero": "乙", "skillSlots": [["乙技一"], ["乙技二"]]},
            {"hero": "丙", "skillSlots": [["丙技一"], ["丙技二"]]},
        ]
        payload = {"formation": "方圆阵", "members": members}
        teams.append(
            {
                "id": make_build_id(payload),
                "ranking": "S",
                "sources": ["strong", "championship"],
                "section": "魏国",
                **payload,
            }
        )
    origin_to_id = {
        f"夺冠御三家!{origin}": teams[index]["id"]
        for index, origin in enumerate(("A5", "F5", "K5"))
    }
    groups = build_championship_groups(
        [["夺冠御三家!A5", "夺冠御三家!F5", "夺冠御三家!K5"]],
        origin_to_id,
    )
    outcomes = [
        ["self" if row == column else "even" for column in range(13)]
        for row in range(13)
    ]
    return {
        "heroes": heroes,
        "skills": skills,
        "formations": {"方圆阵": SQUARE_FORMATION_DESCRIPTION},
        "team": teams,
        "yanwuGuide": {
            "schemaVersion": 1,
            "source": {
                "provider": PROVIDER,
                "workbook": WORKBOOK_NAME,
                "updatedAt": UPDATED_AT,
                "attribution": ATTRIBUTION,
            },
            "matchups": {
                "orientation": "column-build-vs-row-build",
                "buildIds": [team["id"] for team in teams],
                "outcomes": outcomes,
            },
            "championshipGroups": groups,
            "analysisSections": [
                {
                    "section": "魏国",
                    "subject": "甲",
                    "points": ["有效分析"],
                }
            ],
        },
    }


def test_generated_database_invariants_cover_schema_and_references() -> None:
    database = _valid_database()
    validate_generated_database(database)

    unranked_hero = copy.deepcopy(database)
    unranked_hero["heroes"]["甲"].pop("ranking")
    validate_generated_database(unranked_hero)

    broken_reference = copy.deepcopy(database)
    broken_reference["yanwuGuide"]["matchups"]["buildIds"][0] = "missing"
    with pytest.raises(ImportValidationError, match="buildIds"):
        validate_generated_database(broken_reference)

    leaked_contact = copy.deepcopy(database)
    leaked_contact["yanwuGuide"]["analysisSections"][0]["points"] = [
        "VX：850509047"
    ]
    with pytest.raises(ImportValidationError, match="contact"):
        validate_generated_database(leaked_contact)


def test_audited_workbook_cardinalities_fail_closed() -> None:
    valid = ImportStats(
        heroes=100,
        skills=231,
        **EXPECTED_IMPORT_CARDINALITIES,
    )
    validate_import_cardinalities(valid)
    assert valid.ranked_heroes == 99

    missing_build = ImportStats(
        **{
            **valid.__dict__,
            "strong_entries": valid.strong_entries - 1,
            "teams": valid.teams - 1,
        }
    )
    with pytest.raises(ImportValidationError, match="cardinalities changed"):
        validate_import_cardinalities(missing_build)


def test_write_is_default_dry_run_atomic_and_idempotent(tmp_path: Path) -> None:
    database_path = tmp_path / "database.json"
    database_path.write_bytes(b'{"old":true}\n')
    rendered = b'{"new":true}\n'

    assert not write_database_if_changed(
        database_path,
        rendered,
        apply=False,
    )
    assert database_path.read_bytes() == b'{"old":true}\n'

    assert write_database_if_changed(
        database_path,
        rendered,
        apply=True,
    )
    assert database_path.read_bytes() == rendered
    first_stat = database_path.stat()

    assert not write_database_if_changed(
        database_path,
        rendered,
        apply=True,
    )
    second_stat = database_path.stat()
    assert second_stat.st_ino == first_stat.st_ino
    assert second_stat.st_mtime_ns == first_stat.st_mtime_ns
